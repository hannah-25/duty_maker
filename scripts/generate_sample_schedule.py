"""실제 병동 간호사 14명 데이터로 한 달치 스케줄을 생성해 검증 후 엑셀로 저장하는 스크립트.

Streamlit UI(Phase 2~3)가 만들어지기 전, 수간호사가 직접 결과물을 눈으로
확인해볼 수 있도록 하기 위한 임시 확인용 스크립트이다.
생성 결과는 반드시 core/validator.py 전수 검증을 거쳐 위반 리포트와 함께 출력한다.
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.holidays_kr import get_month_holidays
from core.models import (
    ShiftType,
    build_month_requirements,
    compute_month_off_target,
    month_dates,
)
from core.sample_data import build_real_nurses, ward_templates
from core.solver import generate_schedule
from core.validator import validate_schedule

YEAR, MONTH = 2026, 7
WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]

SHIFT_FILLS = {
    "D": PatternFill("solid", fgColor="FFF2CC"),   # 연노랑
    "E": PatternFill("solid", fgColor="DDEBF7"),   # 연파랑
    "N": PatternFill("solid", fgColor="D9D2E9"),   # 연보라
    "S": PatternFill("solid", fgColor="F8CBAD"),   # 연주황
    "O": PatternFill("solid", fgColor="F2F2F2"),   # 연회색
    "연차": PatternFill("solid", fgColor="C6EFCE"),  # 연초록
}
HOLIDAY_HEADER_FILL = PatternFill("solid", fgColor="FFC7A0")  # 주말/공휴일 헤더


def _style_sheet(ws, days, holidays):
    n_days = len(days)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=1 + n_days):
        for cell in row:
            fill = SHIFT_FILLS.get(cell.value)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
    for j, d in enumerate(days):
        header = ws.cell(row=1, column=2 + j)
        header.alignment = Alignment(horizontal="center")
        if d.weekday() >= 5 or d in holidays:
            header.fill = HOLIDAY_HEADER_FILL
            header.font = Font(bold=True, color="C00000")
        ws.column_dimensions[get_column_letter(2 + j)].width = 7
    ws.column_dimensions["A"].width = 10
    for col in range(2 + n_days, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8


def main():
    nurses = build_real_nurses()
    weekday_template, weekend_template = ward_templates()
    requirements = build_month_requirements(YEAR, MONTH, weekday_template, weekend_template)
    holidays = get_month_holidays(YEAR, MONTH)
    off_target_value = compute_month_off_target(YEAR, MONTH, holidays)
    off_target = {n.name: off_target_value for n in nurses}

    result = generate_schedule(nurses, YEAR, MONTH, requirements, off_target, time_limit_seconds=60)
    if not result.feasible:
        print("실행 불가:", result.infeasible_categories)
        return

    report = validate_schedule(
        nurses, YEAR, MONTH, result.assignments, requirements, off_target
    )

    days = month_dates(YEAR, MONTH)
    columns = [f"{d.month}/{d.day}({WEEKDAY_KR[d.weekday()]})" for d in days]
    rows = [
        [result.assignments[(n.name, d)].value for d in days]
        for n in nurses
    ]
    df = pd.DataFrame(rows, columns=columns, index=[n.name for n in nurses])

    per_nurse = report.stats["개인별"]
    for label in ("근무", "N", "O", "연차", "오프편차"):
        df[label] = [per_nurse[n.name][label] for n in nurses]

    out_dir = Path(__file__).resolve().parent.parent / "sample_output"
    out_dir.mkdir(exist_ok=True)

    def save(path):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="듀티표")
            _style_sheet(writer.sheets["듀티표"], days, holidays)

    out_path = out_dir / f"{YEAR}-{MONTH:02d}_sample_schedule.xlsx"
    try:
        save(out_path)
    except PermissionError:
        out_path = out_dir / f"{YEAR}-{MONTH:02d}_sample_schedule_v2.xlsx"
        save(out_path)

    print(f"공휴일(병동 기준): {sorted(holidays)}")
    print(f"목표 오프일수(1인당): {off_target_value}")
    print(f"소프트 벌점: {result.soft_violations}")
    print(report.summary())
    print(f"엑셀 저장 완료: {out_path}")


if __name__ == "__main__":
    main()
