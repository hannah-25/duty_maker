from __future__ import annotations

import io
from datetime import date
from urllib.parse import quote

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Response, status
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from api.deps import CurrentUser, get_current_user, require_admin
from api.routers.schedule import _off_target, _selected_holidays
from api.schemas import ExportSettings
from api.state_store import load_ward_state, save_ward_state
from core.hwpx_export import export_schedule_hwpx
from core.models import ShiftType, month_dates
from core.persistence import list_wards

router = APIRouter(prefix="/api/exports", tags=["exports"])

_DEFAULT_EXPORT_SETTINGS = {
    "title_mode": "ward_month_off",
    "custom_title": "",
    "holiday_color": "#FFE7D8",
    "honored_off_color": "#2563EB",
    "summary_fields": ["E", "N", "O"],
}


def _export_settings(ss: dict) -> dict[str, str]:
    raw = ss.get("export_settings") or {}
    value = dict(_DEFAULT_EXPORT_SETTINGS)
    value.update({key: raw[key] for key in value if key in raw})
    if value["title_mode"] not in {"ward_month_off", "hospital_ward_month_off", "custom"}:
        value["title_mode"] = _DEFAULT_EXPORT_SETTINGS["title_mode"]
    fields = value.get("summary_fields")
    if not isinstance(fields, list):
        fields = _DEFAULT_EXPORT_SETTINGS["summary_fields"]
    value["summary_fields"] = list(dict.fromkeys(field for field in fields if field in {"D", "E", "N", "O", "AL"}))
    for key in ("holiday_color", "honored_off_color"):
        color = str(value[key])
        if not (len(color) == 7 and color.startswith("#") and all(c in "0123456789abcdefABCDEF" for c in color[1:])):
            value[key] = _DEFAULT_EXPORT_SETTINGS[key]
        else:
            value[key] = color.upper()
    return value


def _ensure_visible_schedule(user: CurrentUser) -> dict:
    ss = load_ward_state(user.ward_id)
    result = ss.get("schedule_result")
    if result is None or not result.feasible:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "??? ? ?? ???? ????.")
    if not user.is_admin and not ss.get("result_published", False):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "??? ???? ????.")
    return ss


def _request_highlight_days(result) -> set[tuple[str, date]]:
    return {
        (req.nurse_name, req.day)
        for req in result.honored_duty_requests
        if getattr(req, "kind", "prefer") == "prefer"
        and req.requested_shift in (ShiftType.O, ShiftType.AL)
    }


def _assistant_marks(ss: dict) -> dict[tuple[str, date], ShiftType]:
    names = {assistant.name for assistant in ss.get("assistants", [])}
    return {
        (req.nurse_name, req.day): req.requested_shift
        for req in ss.get("duty_requests", [])
        if req.nurse_name in names
        and getattr(req, "decision", "force") != "ignore"
        and getattr(req, "kind", "prefer") == "prefer"
    }


def _download_response(data: bytes, filename: str, media_type: str) -> Response:
    quoted = quote(filename)
    return Response(content=data, media_type=media_type, headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quoted}"})


def _title(user: CurrentUser, settings: dict[str, str], month: int, off_target_value: int) -> str:
    ward = list_wards().get(user.ward_id, {})
    ward_name = ward.get("ward_name", "\ubcd1\ub3d9")
    hospital_name = ward.get("hospital_name", "")
    if settings["title_mode"] == "custom" and settings["custom_title"].strip():
        prefix = settings["custom_title"].strip()
    elif settings["title_mode"] == "hospital_ward_month_off" and hospital_name:
        prefix = f"{hospital_name} {ward_name}"
    else:
        prefix = ward_name
    return f"{prefix} {month}\uc6d4 \uadfc\ubb34\ud45c (OFF {off_target_value}\uac1c)"


@router.get("/settings", response_model=ExportSettings)
def get_export_settings(user: CurrentUser = Depends(get_current_user)) -> ExportSettings:
    return ExportSettings(**_export_settings(load_ward_state(user.ward_id)))


@router.put("/settings", response_model=ExportSettings)
def put_export_settings(body: ExportSettings, user: CurrentUser = Depends(require_admin)) -> ExportSettings:
    ss = load_ward_state(user.ward_id)
    ss["export_settings"] = body.model_dump()
    save_ward_state(user.ward_id, ss)
    return ExportSettings(**_export_settings(ss))


@router.get("/hwpx")
def export_hwpx(user: CurrentUser = Depends(get_current_user)) -> Response:
    ss = _ensure_visible_schedule(user)
    year, month = int(ss.get("year", 2026)), int(ss.get("month", 7))
    holidays = _selected_holidays(ss, year, month)
    off_target_values = _off_target(ss, year, month)
    off_target_value = next(iter(off_target_values.values()), 0)
    result = ss["schedule_result"]
    settings = _export_settings(ss)
    data = export_schedule_hwpx(
        ss.get("nurses", []), year, month, result.assignments, holidays, off_target_value,
        _request_highlight_days(result), assistants=ss.get("assistants", []), assistant_marks=_assistant_marks(ss),
        title=_title(user, settings, month, off_target_value), holiday_color=settings["holiday_color"],
        honored_off_color=settings["honored_off_color"], summary_fields=settings["summary_fields"],
    )
    return _download_response(data, f"{year}-{month:02d}_duty_schedule.hwpx", "application/octet-stream")


@router.get("/xlsx")
def export_xlsx(user: CurrentUser = Depends(get_current_user)) -> Response:
    ss = _ensure_visible_schedule(user)
    year, month = int(ss.get("year", 2026)), int(ss.get("month", 7))
    result = ss["schedule_result"]
    days = month_dates(year, month)
    holidays = _selected_holidays(ss, year, month)
    off_target_value = next(iter(_off_target(ss, year, month).values()), 0)
    settings = _export_settings(ss)
    nurses = ss.get("nurses", [])
    frame = pd.DataFrame(
        [[result.assignments[(nurse.name, day)].value for day in days] for nurse in nurses],
        index=[nurse.name for nurse in nurses], columns=[f"{day.month}/{day.day}" for day in days],
    )
    summary_labels = {"D": "D", "E": "E", "N": "N", "O": "O", "AL": "\uc5f0\ucc28"}
    for field in settings["summary_fields"]:
        shift = ShiftType.AL if field == "AL" else ShiftType(field)
        frame[summary_labels[field]] = [
            sum(result.assignments[(nurse.name, day)] == shift for day in days)
            for nurse in nurses
        ]
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="\uadfc\ubb34\ud45c", startrow=2)
        sheet = writer.book["\uadfc\ubb34\ud45c"]
        last_column = len(days) + 1
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
        sheet.cell(1, 1).value = _title(user, settings, month, off_target_value)
        sheet.cell(1, 1).font = Font(bold=True, size=14)
        holiday_fill = PatternFill("solid", fgColor=settings["holiday_color"].lstrip("#"))
        requested_font = Font(color=settings["honored_off_color"].lstrip("#"))
        highlighted = _request_highlight_days(result)
        for index, day in enumerate(days, start=2):
            if day.weekday() >= 5 or day in holidays:
                for row in range(3, len(nurses) + 4):
                    sheet.cell(row, index).fill = holiday_fill
        for row, nurse in enumerate(nurses, start=4):
            for column, day in enumerate(days, start=2):
                if (nurse.name, day) in highlighted:
                    sheet.cell(row, column).font = requested_font
        sheet.freeze_panes = "B4"
        sheet.column_dimensions["A"].width = 14
        for column in range(2, last_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 5
    return _download_response(out.getvalue(), f"{year}-{month:02d}_duty_schedule.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
